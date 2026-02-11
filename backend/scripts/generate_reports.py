"""Report Generation Script

Generates formatted Excel reports and PDF summaries for insurance data.
Demonstrates openpyxl styling, data aggregation, and automated report
distribution (email simulation).

Usage:
    python scripts/generate_reports.py --type [premiums|claims|products] [--email]
"""
import os
import sys
import argparse
import logging
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app, db
from app.models import Insurer, Product, Policy, Claim

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from sqlalchemy import func

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Common styles
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
TITLE_FONT = Font(bold=True, size=16, color='1B4F72')
SUBTITLE_FONT = Font(italic=True, size=10, color='666666')
CURRENCY_FORMAT = '€#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def write_header_row(ws, row, headers):
    """Write a styled header row to a worksheet."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)


def generate_premium_report(output_dir):
    """Generate a detailed premium report with charts."""
    logger.info('Generating premium report...')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Premium Overzicht'

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'InsuranceHub - Maandelijks Premium Rapport'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Rapport datum: {datetime.now().strftime("%d %B %Y")}'
    ws['A2'].font = SUBTITLE_FONT

    # Summary section
    total_premium = db.session.query(func.sum(Policy.premium_amount)).filter(
        Policy.status == 'active'
    ).scalar() or 0
    total_policies = Policy.query.filter(Policy.status == 'active').count()

    ws['A4'] = 'Samenvatting'
    ws['A4'].font = Font(bold=True, size=12, color='1B4F72')
    ws['A5'] = 'Totaal actieve polissen:'
    ws['B5'] = total_policies
    ws['A6'] = 'Totale maandpremie:'
    ws['B6'] = float(total_premium)
    ws['B6'].number_format = CURRENCY_FORMAT
    ws['A7'] = 'Gemiddelde premie per polis:'
    ws['B7'] = round(float(total_premium) / max(total_policies, 1), 2)
    ws['B7'].number_format = CURRENCY_FORMAT

    # Detail table: Premium per verzekeraar
    headers = ['Verzekeraar', 'Product Type', 'Aantal Polissen',
               'Totale Premie', 'Gemiddelde Premie', 'Min. Premie', 'Max. Premie']
    write_header_row(ws, 9, headers)

    results = db.session.query(
        Insurer.name, Product.product_type,
        func.count(Policy.id), func.sum(Policy.premium_amount),
        func.avg(Policy.premium_amount), func.min(Policy.premium_amount),
        func.max(Policy.premium_amount),
    ).join(Product, Product.insurer_id == Insurer.id)\
     .join(Policy, Policy.product_id == Product.id)\
     .filter(Policy.status == 'active')\
     .group_by(Insurer.name, Product.product_type)\
     .order_by(Insurer.name)\
     .all()

    for row_idx, r in enumerate(results, 10):
        ws.cell(row=row_idx, column=1, value=r[0]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=r[1]).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=r[2]).border = THIN_BORDER
        for col in range(4, 8):
            cell = ws.cell(row=row_idx, column=col,
                          value=round(float(r[col - 1] or 0), 2))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER

    # Add pie chart: Premium per insurer
    chart_ws = wb.create_sheet('Grafieken')
    chart_ws['A1'] = 'Premium Verdeling per Verzekeraar'
    chart_ws['A1'].font = TITLE_FONT

    insurer_premiums = db.session.query(
        Insurer.name, func.sum(Policy.premium_amount)
    ).join(Product, Product.insurer_id == Insurer.id)\
     .join(Policy, Policy.product_id == Product.id)\
     .filter(Policy.status == 'active')\
     .group_by(Insurer.name)\
     .all()

    for i, (name, premium) in enumerate(insurer_premiums, 3):
        chart_ws.cell(row=i, column=1, value=name)
        chart_ws.cell(row=i, column=2, value=float(premium))

    if insurer_premiums:
        pie = PieChart()
        pie.title = 'Premie Verdeling per Verzekeraar'
        labels = Reference(chart_ws, min_col=1, min_row=3,
                          max_row=2 + len(insurer_premiums))
        data = Reference(chart_ws, min_col=2, min_row=3,
                        max_row=2 + len(insurer_premiums))
        pie.add_data(data)
        pie.set_categories(labels)
        pie.width = 20
        pie.height = 15
        chart_ws.add_chart(pie, 'D3')

    auto_fit_columns(ws)

    filepath = os.path.join(output_dir, f'premium_report_{datetime.now().strftime("%Y%m%d")}.xlsx')
    wb.save(filepath)
    logger.info(f'Premium report saved to {filepath}')
    return filepath


def generate_claims_report(output_dir):
    """Generate a claims analysis report."""
    logger.info('Generating claims report...')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Schade Analyse'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'InsuranceHub - Schade Analyse Rapport'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Rapport datum: {datetime.now().strftime("%d %B %Y")}'
    ws['A2'].font = SUBTITLE_FONT

    # Claims overview
    total_claims = Claim.query.count()
    total_claimed = db.session.query(func.sum(Claim.amount)).scalar() or 0
    total_approved = db.session.query(func.sum(Claim.approved_amount)).scalar() or 0

    ws['A4'] = 'Samenvatting'
    ws['A4'].font = Font(bold=True, size=12, color='1B4F72')
    ws['A5'] = 'Totaal claims:'
    ws['B5'] = total_claims
    ws['A6'] = 'Totaal geclaimd:'
    ws['B6'] = float(total_claimed)
    ws['B6'].number_format = CURRENCY_FORMAT
    ws['A7'] = 'Totaal goedgekeurd:'
    ws['B7'] = float(total_approved)
    ws['B7'].number_format = CURRENCY_FORMAT

    # Detail table
    headers = ['Claim Nr.', 'Polis Nr.', 'Klant', 'Categorie',
               'Bedrag', 'Goedgekeurd', 'Status']
    write_header_row(ws, 9, headers)

    claims = Claim.query.order_by(Claim.claim_date.desc()).all()
    for row_idx, claim in enumerate(claims, 10):
        ws.cell(row=row_idx, column=1, value=claim.claim_number).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=claim.policy.policy_number if claim.policy else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=claim.policy.customer_name if claim.policy else '').border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=claim.category or '').border = THIN_BORDER

        amount_cell = ws.cell(row=row_idx, column=5, value=float(claim.amount or 0))
        amount_cell.number_format = CURRENCY_FORMAT
        amount_cell.border = THIN_BORDER

        approved_cell = ws.cell(row=row_idx, column=6, value=float(claim.approved_amount or 0))
        approved_cell.number_format = CURRENCY_FORMAT
        approved_cell.border = THIN_BORDER

        status_cell = ws.cell(row=row_idx, column=7, value=claim.status)
        status_cell.border = THIN_BORDER
        # Color-code status
        status_colors = {
            'paid': '27AE60', 'approved': '2ECC71',
            'under_review': 'F39C12', 'submitted': '3498DB',
            'rejected': 'E74C3C',
        }
        if claim.status in status_colors:
            status_cell.fill = PatternFill(
                start_color=status_colors[claim.status],
                end_color=status_colors[claim.status],
                fill_type='solid'
            )
            status_cell.font = Font(color='FFFFFF', bold=True)

    auto_fit_columns(ws)

    filepath = os.path.join(output_dir, f'claims_report_{datetime.now().strftime("%Y%m%d")}.xlsx')
    wb.save(filepath)
    logger.info(f'Claims report saved to {filepath}')
    return filepath


def generate_product_report(output_dir):
    """Generate a product catalog comparison report."""
    logger.info('Generating product report...')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product Catalogus'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'InsuranceHub - Product Vergelijking'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Rapport datum: {datetime.now().strftime("%d %B %Y")}'
    ws['A2'].font = SUBTITLE_FONT

    headers = ['Verzekeraar', 'Product Code', 'Product Naam', 'Type',
               'Basispremie', 'Dekking', 'Eigen Risico', 'Verkochte Polissen']
    write_header_row(ws, 4, headers)

    results = db.session.query(
        Insurer.name, Product.product_code, Product.name, Product.product_type,
        Product.base_premium, Product.coverage_amount, Product.deductible,
        func.count(Policy.id),
    ).join(Product, Product.insurer_id == Insurer.id)\
     .outerjoin(Policy, Policy.product_id == Product.id)\
     .group_by(Insurer.name, Product.product_code, Product.name,
               Product.product_type, Product.base_premium,
               Product.coverage_amount, Product.deductible)\
     .order_by(Insurer.name, Product.product_type)\
     .all()

    for row_idx, r in enumerate(results, 5):
        ws.cell(row=row_idx, column=1, value=r[0]).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=r[1]).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=r[2]).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=r[3]).border = THIN_BORDER
        for col, val in [(5, r[4]), (6, r[5]), (7, r[6])]:
            cell = ws.cell(row=row_idx, column=col, value=float(val or 0))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
        ws.cell(row=row_idx, column=8, value=r[7]).border = THIN_BORDER

    auto_fit_columns(ws)

    filepath = os.path.join(output_dir, f'product_report_{datetime.now().strftime("%Y%m%d")}.xlsx')
    wb.save(filepath)
    logger.info(f'Product report saved to {filepath}')
    return filepath


def simulate_email(filepath, report_type):
    """Simulate sending report via email (logs the action)."""
    logger.info(f'[EMAIL SIMULATION] Sending {report_type} report')
    logger.info(f'  To: management@insurancehub.nl')
    logger.info(f'  Subject: InsuranceHub {report_type.title()} Rapport - {datetime.now().strftime("%d-%m-%Y")}')
    logger.info(f'  Attachment: {os.path.basename(filepath)}')
    logger.info(f'  Status: Sent (simulated)')


def main():
    parser = argparse.ArgumentParser(description='Generate insurance reports')
    parser.add_argument('--type', '-t', choices=['premiums', 'claims', 'products', 'all'],
                       default='all', help='Report type to generate')
    parser.add_argument('--output', '-o', default='reports',
                       help='Output directory for reports')
    parser.add_argument('--email', '-e', action='store_true',
                       help='Simulate email distribution')
    args = parser.parse_args()

    app = create_app()
    with app.app_context():
        os.makedirs(args.output, exist_ok=True)

        report_generators = {
            'premiums': generate_premium_report,
            'claims': generate_claims_report,
            'products': generate_product_report,
        }

        types = report_generators.keys() if args.type == 'all' else [args.type]

        for report_type in types:
            filepath = report_generators[report_type](args.output)
            if args.email:
                simulate_email(filepath, report_type)

        logger.info('\nAll reports generated successfully!')


if __name__ == '__main__':
    main()
