"""Generate Excel Import Template

Creates a formatted Excel template for importing insurance products,
with data validation rules, example data, and an instructions sheet.

The generated template includes:
- Pre-formatted columns with headers
- Data validation dropdowns for product type and insurer code
- Example data rows
- Instructions sheet explaining each field
- Conditional formatting for required fields

Usage:
    python scripts/create_excel_template.py [--output path.xlsx]
"""
import os
import sys
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def create_template(output_path):
    """Create the product import Excel template."""
    wb = openpyxl.Workbook()

    # ---- Instructions Sheet ----
    ws_inst = wb.active
    ws_inst.title = 'Instructies'
    ws_inst.sheet_properties.tabColor = '1B4F72'

    title_font = Font(bold=True, size=16, color='1B4F72')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws_inst.merge_cells('A1:F1')
    ws_inst['A1'] = 'InsuranceHub - Product Import Template'
    ws_inst['A1'].font = title_font

    ws_inst['A3'] = 'Instructies voor gebruik:'
    ws_inst['A3'].font = Font(bold=True, size=12)

    instructions = [
        '1. Ga naar het tabblad "Products" om productgegevens in te voeren',
        '2. Vul alle vereiste velden in (gemarkeerd met *)',
        '3. Gebruik de dropdown menu\'s voor Product Type en Verzekeraar Code',
        '4. Premies en bedragen moeten positieve getallen zijn',
        '5. Product codes moeten uniek zijn en het formaat XXX-YYY-NNN volgen',
        '6. Sla het bestand op als .xlsx en upload via InsuranceHub > Products > Import',
        '',
        'Veld beschrijvingen:',
    ]
    for i, text in enumerate(instructions, 4):
        ws_inst[f'A{i}'] = text
        ws_inst[f'A{i}'].font = Font(size=10)

    # Field descriptions table
    fields = [
        ('product_code *', 'Unieke productcode (bijv. ACH-WOZ-001)', 'Tekst'),
        ('name *', 'Productnaam in het Nederlands', 'Tekst'),
        ('product_type *', 'Type verzekering', 'Keuze: property, life, health, auto, liability, travel'),
        ('insurer_code *', 'Code van de verzekeraar', 'Keuze: ACH, AEG, ALZ, ASR, NNG, NAT'),
        ('base_premium *', 'Maandelijkse basispremie in EUR', 'Getal (positief)'),
        ('coverage_amount', 'Maximale dekking in EUR', 'Getal (positief)'),
        ('deductible', 'Eigen risico in EUR', 'Getal (0 of positief)'),
        ('description', 'Korte beschrijving van het product', 'Tekst'),
    ]

    row = 13
    for col, header in enumerate(['Veld', 'Beschrijving', 'Formaat'], 1):
        cell = ws_inst.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for i, (field, desc, fmt) in enumerate(fields, row + 1):
        ws_inst.cell(row=i, column=1, value=field).border = border
        ws_inst.cell(row=i, column=2, value=desc).border = border
        ws_inst.cell(row=i, column=3, value=fmt).border = border
        if '*' in field:
            ws_inst.cell(row=i, column=1).font = Font(bold=True, color='CC0000')

    for col in ['A', 'B', 'C']:
        ws_inst.column_dimensions[col].width = 35

    # ---- Products Sheet ----
    ws = wb.create_sheet('Products')
    ws.sheet_properties.tabColor = '27AE60'

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Product Import - Vul de gegevens hieronder in'
    ws['A1'].font = Font(bold=True, size=13, color='1B4F72')

    # Headers
    headers = [
        ('product_code', 15),
        ('name', 35),
        ('product_type', 15),
        ('insurer_code', 14),
        ('base_premium', 14),
        ('coverage_amount', 16),
        ('deductible', 12),
        ('description', 40),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data validation: product type dropdown
    type_validation = DataValidation(
        type='list',
        formula1='"property,life,health,auto,liability,travel"',
        allow_blank=False,
    )
    type_validation.error = 'Kies een geldig producttype'
    type_validation.errorTitle = 'Ongeldig type'
    type_validation.prompt = 'Kies het producttype'
    type_validation.promptTitle = 'Product Type'
    ws.add_data_validation(type_validation)
    type_validation.add(f'C4:C100')

    # Data validation: insurer code dropdown
    insurer_validation = DataValidation(
        type='list',
        formula1='"ACH,AEG,ALZ,ASR,NNG,NAT"',
        allow_blank=False,
    )
    insurer_validation.error = 'Kies een geldige verzekeraar code'
    insurer_validation.errorTitle = 'Ongeldige verzekeraar'
    ws.add_data_validation(insurer_validation)
    insurer_validation.add(f'D4:D100')

    # Data validation: positive numbers for premium
    premium_validation = DataValidation(
        type='decimal',
        operator='greaterThan',
        formula1='0',
        allow_blank=False,
    )
    premium_validation.error = 'Premie moet een positief getal zijn'
    ws.add_data_validation(premium_validation)
    premium_validation.add(f'E4:E100')

    # Example data rows
    examples = [
        ('ACH-WOZ-003', 'Woonhuis Verzekering Premium', 'property', 'ACH', 55.00, 750000, 100, 'Premium woonhuisverzekering met uitgebreide dekking'),
        ('AEG-AUTO-001', 'Autoverzekering Compleet', 'auto', 'AEG', 65.00, 2500000, 350, 'Complete autoverzekering met pechhulp'),
        ('ALZ-REIS-001', 'Wereldreisverzekering', 'travel', 'ALZ', 8.50, 75000, 50, 'Doorlopende reisverzekering wereldwijd'),
    ]

    example_fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    for row_idx, example in enumerate(examples, 4):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.border = border
            if col_idx in (5, 6, 7):
                cell.number_format = '#,##0.00'

    # Mark example rows
    ws.cell(row=4, column=9, value='<-- Voorbeeldrijen (verwijderen voor import)').font = Font(italic=True, color='999999')

    # Freeze header row
    ws.freeze_panes = 'A4'

    wb.save(output_path)
    print(f'Template saved to: {output_path}')
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Generate Excel import template')
    parser.add_argument('--output', '-o',
                       default=os.path.join('..', 'sample-data', 'excel-templates', 'product-import-template.xlsx'),
                       help='Output file path')
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    create_template(args.output)


if __name__ == '__main__':
    main()
