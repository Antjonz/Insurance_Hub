"""Excel handler for insurance product and policy data.

Provides functionality to:
- Read .xlsx files with products/policies (import)
- Export data to formatted Excel workbooks (export)
- Support different sheet structures from various insurers
- Data cleaning and normalization
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


# Expected column mappings for product import sheets
PRODUCT_COLUMN_MAPPINGS = {
    'product_code': ['product_code', 'productcode', 'code', 'sku', 'product id'],
    'name': ['name', 'product_name', 'productnaam', 'titel', 'title'],
    'product_type': ['product_type', 'type', 'category', 'producttype', 'soort'],
    'base_premium': ['base_premium', 'premium', 'premie', 'maandpremie', 'monthly_premium'],
    'coverage_amount': ['coverage_amount', 'coverage', 'dekking', 'dekkingsbedrag', 'sum_insured'],
    'deductible': ['deductible', 'excess', 'eigen_risico', 'eigenrisico', 'own_risk'],
    'description': ['description', 'beschrijving', 'omschrijving', 'details'],
    'insurer_code': ['insurer_code', 'verzekeraar_code', 'insurer', 'verzekeraar'],
}


def detect_columns(header_row: list) -> dict:
    """Map actual column headers to standard field names.

    Handles Dutch and English column names, case-insensitive matching.
    """
    column_map = {}
    for col_idx, header in enumerate(header_row):
        if not header:
            continue
        header_lower = str(header).lower().strip().replace(' ', '_')

        for standard_field, possible_names in PRODUCT_COLUMN_MAPPINGS.items():
            if header_lower in possible_names:
                column_map[standard_field] = col_idx
                break

    return column_map


def parse_excel_products(filepath: str) -> list:
    """Parse an Excel file containing insurance product data.

    Supports flexible column ordering and naming (Dutch/English).
    Automatically detects the header row and maps columns.

    Args:
        filepath: Path to the .xlsx file

    Returns:
        List of normalized product dictionaries
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    products = []

    # Try each sheet - look for product data
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # Detect header row (first row with recognizable column names)
        header_row_idx = None
        column_map = {}
        for idx, row in enumerate(rows[:5]):  # Check first 5 rows for header
            possible_map = detect_columns(list(row))
            if len(possible_map) >= 2:  # At least 2 recognized columns
                header_row_idx = idx
                column_map = possible_map
                break

        if header_row_idx is None:
            continue

        # Parse data rows
        for row in rows[header_row_idx + 1:]:
            if not any(row):  # Skip empty rows
                continue

            product = {}
            for field, col_idx in column_map.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    value = row[col_idx]
                    # Convert numeric fields
                    if field in ('base_premium', 'coverage_amount', 'deductible'):
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            value = 0
                    else:
                        value = str(value).strip()
                    product[field] = value

            # Only add if we have the essential fields
            if product.get('product_code') and product.get('name'):
                # Extract insurer code from product_code if not in data
                if 'insurer_code' not in product and '-' in product.get('product_code', ''):
                    product['insurer_code'] = product['product_code'].split('-')[0]
                products.append(product)

    wb.close()
    return products


def create_product_export(products: list, filepath: str):
    """Export product data to a formatted Excel file.

    Creates a professional-looking spreadsheet with:
    - Styled header row
    - Formatted currency columns
    - Auto-fitted column widths
    - Summary row at the bottom
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    currency_format = '€#,##0.00'

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'InsuranceHub - Product Export'
    ws['A1'].font = Font(bold=True, size=14, color='1B4F72')
    ws['A2'] = f'Exported: {datetime.now().strftime("%d-%m-%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, color='666666')

    # Headers
    headers = ['Product Code', 'Name', 'Insurer', 'Type', 'Base Premium',
               'Coverage', 'Deductible', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data rows
    for row_idx, product in enumerate(products, 5):
        ws.cell(row=row_idx, column=1, value=product.get('product_code', '')).border = border
        ws.cell(row=row_idx, column=2, value=product.get('name', '')).border = border
        ws.cell(row=row_idx, column=3, value=product.get('insurer_name', '')).border = border
        ws.cell(row=row_idx, column=4, value=product.get('product_type', '')).border = border

        for col, field in [(5, 'base_premium'), (6, 'coverage_amount'), (7, 'deductible')]:
            cell = ws.cell(row=row_idx, column=col,
                          value=product.get(field, 0))
            cell.number_format = currency_format
            cell.border = border

        ws.cell(row=row_idx, column=8, value=product.get('status', 'active')).border = border

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 35)

    wb.save(filepath)
    return filepath
