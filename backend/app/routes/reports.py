"""Report routes - generate and export insurance reports."""
import io
from flask import Blueprint, jsonify, request, send_file
from sqlalchemy import func
from datetime import datetime
from app import db
from app.models import Insurer, Product, Policy, Claim

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports/premiums')
def premium_report():
    """Generate a premium report grouped by insurer and product type.

    Demonstrates complex SQL aggregation queries across multiple tables.
    """
    # Date range filter
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = db.session.query(
        Insurer.name.label('insurer_name'),
        Product.product_type,
        func.count(Policy.id).label('policy_count'),
        func.sum(Policy.premium_amount).label('total_premium'),
        func.avg(Policy.premium_amount).label('avg_premium'),
        func.min(Policy.premium_amount).label('min_premium'),
        func.max(Policy.premium_amount).label('max_premium'),
    ).join(Product, Product.insurer_id == Insurer.id)\
     .join(Policy, Policy.product_id == Product.id)\
     .filter(Policy.status == 'active')

    if start_date:
        query = query.filter(Policy.start_date >= start_date)
    if end_date:
        query = query.filter(Policy.start_date <= end_date)

    results = query.group_by(Insurer.name, Product.product_type)\
                   .order_by(Insurer.name, Product.product_type)\
                   .all()

    return jsonify({
        'report_type': 'premium_summary',
        'generated_at': datetime.now().isoformat(),
        'data': [
            {
                'insurer_name': r.insurer_name,
                'product_type': r.product_type,
                'policy_count': r.policy_count,
                'total_premium': float(r.total_premium),
                'avg_premium': round(float(r.avg_premium), 2),
                'min_premium': float(r.min_premium),
                'max_premium': float(r.max_premium),
            }
            for r in results
        ],
    })


@reports_bp.route('/reports/claims')
def claims_report():
    """Generate claims analysis report."""
    query = db.session.query(
        Insurer.name.label('insurer_name'),
        Claim.status,
        Claim.category,
        func.count(Claim.id).label('claim_count'),
        func.sum(Claim.amount).label('total_claimed'),
        func.sum(Claim.approved_amount).label('total_approved'),
    ).join(Policy, Claim.policy_id == Policy.id)\
     .join(Product, Policy.product_id == Product.id)\
     .join(Insurer, Product.insurer_id == Insurer.id)

    results = query.group_by(Insurer.name, Claim.status, Claim.category)\
                   .order_by(Insurer.name)\
                   .all()

    return jsonify({
        'report_type': 'claims_analysis',
        'generated_at': datetime.now().isoformat(),
        'data': [
            {
                'insurer_name': r.insurer_name,
                'status': r.status,
                'category': r.category,
                'claim_count': r.claim_count,
                'total_claimed': float(r.total_claimed) if r.total_claimed else 0,
                'total_approved': float(r.total_approved) if r.total_approved else 0,
            }
            for r in results
        ],
    })


@reports_bp.route('/reports/products')
def products_report():
    """Generate product performance comparison report."""
    results = db.session.query(
        Insurer.name.label('insurer_name'),
        Product.name.label('product_name'),
        Product.product_type,
        Product.base_premium,
        Product.coverage_amount,
        func.count(Policy.id).label('policies_sold'),
        func.sum(Policy.premium_amount).label('revenue'),
    ).join(Product, Product.insurer_id == Insurer.id)\
     .outerjoin(Policy, Policy.product_id == Product.id)\
     .group_by(Insurer.name, Product.name, Product.product_type,
               Product.base_premium, Product.coverage_amount)\
     .order_by(func.count(Policy.id).desc())\
     .all()

    return jsonify({
        'report_type': 'product_performance',
        'generated_at': datetime.now().isoformat(),
        'data': [
            {
                'insurer_name': r.insurer_name,
                'product_name': r.product_name,
                'product_type': r.product_type,
                'base_premium': float(r.base_premium),
                'coverage_amount': float(r.coverage_amount) if r.coverage_amount else None,
                'policies_sold': r.policies_sold,
                'revenue': float(r.revenue) if r.revenue else 0,
            }
            for r in results
        ],
    })


@reports_bp.route('/reports/export', methods=['POST'])
def export_report():
    """Export a report to Excel format.

    Demonstrates Excel generation with openpyxl, including:
    - Formatted headers with styling
    - Data type formatting (currency, dates)
    - Auto-fitted column widths
    - Summary rows
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json()
    report_type = data.get('report_type', 'premiums')

    wb = openpyxl.Workbook()
    ws = wb.active

    # Styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    if report_type == 'premiums':
        ws.title = 'Premium Report'

        # Title row
        ws.merge_cells('A1:G1')
        ws['A1'] = 'InsuranceHub - Premium Rapport'
        ws['A1'].font = Font(bold=True, size=14, color='1B4F72')
        ws['A2'] = f'Gegenereerd: {datetime.now().strftime("%d-%m-%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, color='666666')

        headers = ['Verzekeraar', 'Producttype', 'Aantal Polissen',
                   'Totale Premie', 'Gem. Premie', 'Min. Premie', 'Max. Premie']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Query data
        results = db.session.query(
            Insurer.name, Product.product_type,
            func.count(Policy.id), func.sum(Policy.premium_amount),
            func.avg(Policy.premium_amount), func.min(Policy.premium_amount),
            func.max(Policy.premium_amount),
        ).join(Product, Product.insurer_id == Insurer.id)\
         .join(Policy, Policy.product_id == Product.id)\
         .filter(Policy.status == 'active')\
         .group_by(Insurer.name, Product.product_type)\
         .order_by(Insurer.name)\
         .all()

        for row_idx, r in enumerate(results, 5):
            ws.cell(row=row_idx, column=1, value=r[0]).border = border
            ws.cell(row=row_idx, column=2, value=r[1]).border = border
            ws.cell(row=row_idx, column=3, value=r[2]).border = border
            for col in range(4, 8):
                cell = ws.cell(row=row_idx, column=col,
                              value=round(float(r[col - 1]), 2) if r[col - 1] else 0)
                cell.number_format = '€#,##0.00'
                cell.border = border

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    elif report_type == 'claims':
        ws.title = 'Claims Report'
        ws.merge_cells('A1:F1')
        ws['A1'] = 'InsuranceHub - Schade Rapport'
        ws['A1'].font = Font(bold=True, size=14, color='1B4F72')

        headers = ['Verzekeraar', 'Status', 'Categorie',
                   'Aantal Claims', 'Totaal Geclaimd', 'Totaal Goedgekeurd']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        results = db.session.query(
            Insurer.name, Claim.status, Claim.category,
            func.count(Claim.id), func.sum(Claim.amount),
            func.sum(Claim.approved_amount),
        ).join(Policy, Claim.policy_id == Policy.id)\
         .join(Product, Policy.product_id == Product.id)\
         .join(Insurer, Product.insurer_id == Insurer.id)\
         .group_by(Insurer.name, Claim.status, Claim.category)\
         .all()

        for row_idx, r in enumerate(results, 5):
            ws.cell(row=row_idx, column=1, value=r[0]).border = border
            ws.cell(row=row_idx, column=2, value=r[1]).border = border
            ws.cell(row=row_idx, column=3, value=r[2]).border = border
            ws.cell(row=row_idx, column=4, value=r[3]).border = border
            for col in range(5, 7):
                cell = ws.cell(row=row_idx, column=col,
                              value=round(float(r[col - 1]), 2) if r[col - 1] else 0)
                cell.number_format = '€#,##0.00'
                cell.border = border

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'insurancehub_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
